#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#@time:
#@Author:lsy
#@file:
#@function:-----------

from .models import HostInfoManage,UserManage
from django.http import HttpResponse
from io import StringIO
from automatization.settings import logger
from threading import Thread
from django.core.paginator import PageNotAnInteger,Paginator,EmptyPage
from .key_decode import AESEBC
import numpy as np
import copy
import paramiko
import socket
import json
import os
import xlrd
import xlwt
import openpyxl

# 定义一个的线程
def Tr_async(fun):
    def wrapper(*args,**kwargs):
        thr = Thread(target=fun,args=args,kwargs=kwargs)
        thr.start()
    return wrapper


#插入主机信息
def Ins_HostInfo(request):
    try:
        logger.info('<---调用主机添加接口--->')
        body = json.loads(request.body.decode())
        logger.info('<---接收的参数:{0}--->'.format(body))

        hostname = body.get('hostname','')
        hostgroup = body.get('hostgroup','')
        hostip = body.get('hostip','')
        port = body.get('port',0)
        credential = body.get('credential','')
        description = body.get('description','')



        hostinfomanage = HostInfoManage()
        hostinfomanage.hostname = hostname
        hostinfomanage.hostgroup = hostgroup
        hostinfomanage.hostip = hostip
        hostinfomanage.port = port
        hostinfomanage.credential = credential
        hostinfomanage.description = description
        # hostinfomanage.hoststatus = hoststatus
        hostinfomanage.save()

        logger.info('<---主机信息添加成功--->'.format(body))
        data = {'code':1,'msg':'主机信息添加成功','data':'success'}
        return HttpResponse(json.dumps(data),content_type='application/json')
    except Exception as e:
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return HttpResponse(json.dumps(data),content_type='application/json')


# 单笔主机信息连通性校验
def  Single_Connect_Check(request):
    logger.info('<---调用单笔主机信息连通性校验接口--->')
    try:
        body = json.loads(request.body.decode())
        logger.info('<---获取参数:{0}--->'.format(body))
        hostip = body.get('hostip','')
        port = body.get('port','')
        credential = body.get('credential','')
        if hostip and credential and port:
            #调用连通性接口
            Clint_Host_Test([{'hostip':hostip,'port':port,'credential':credential}])

            logger.info('<---连通性校验启动完成 请查看数据库状态码--->')
            data = {'code':1,'msg':'连通性校验启动完成 请查看数据库状态码','data':'success'}
            return HttpResponse(json.dumps(data),content_type='application/json')
        else:
            logger.info('<---主机ip 端口 访问凭证存在空值 ip:{0} port:{1} 凭证:{2}--->'.format(hostip,port,credential))
            data = {'code':1,'msg':'主机信息添加成功','data':'success'}
            return HttpResponse(json.dumps(data),content_type='application/json')



    except Exception as e:
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))

#批量插入
def Bulk_Create_HostInfo(request):
    try:
        logger.info('<---调用批量插入主机信息接口--->')
        body = json.loads(request.body.decode())
        logger.info('<---接收的参数:{0}--->'.format(body))

        filepath = body.get('filepath','')
        short_name, extension = os.path.splitext(filepath)
        if extension == '.xlsx':
            sheet_host_list = XLSX_Read_Excel(filepath)
        else:
            sheet_host_list = XLS_Read_Excel(filepath)
        #存储插入主机表的存储
        bulk_create_host_list = []
        #存储插入主机用户表的存储
        bulk_create_user_list = []
        #存储ip和端口 为后续连通性测试准备
        clint_host_list = []
        clint_host_dict = {}
        for sheet_value_list in sheet_host_list:
            #存ip和端口
            clint_host_dict["hostip"] = sheet_value_list["hostip"]
            clint_host_dict["port"] = sheet_value_list["port"]
            clint_host_dict["credential"] = sheet_value_list["credential"]
            #插入主机对象
            bulk_create_host_value = HostInfoManage(hostname=sheet_value_list["hostname"],hostgroup=str(sheet_value_list["hostgroup"]),hostip=sheet_value_list["hostip"],
                                                    port=sheet_value_list["port"],credential=str(sheet_value_list["credential"]),description=sheet_value_list["description"])
            bulk_create_host_list.append(bulk_create_host_value)

            #插入主机用户对象
            #密码加密
            aes_password = sheet_value_list["password"]
            aes_become_password = sheet_value_list["become_password"]
            if aes_password:
                aes_password = AESEBC().Encrypt(aes_password)

            if aes_become_password:
                aes_become_password = AESEBC().Encrypt(aes_become_password)

            bulk_create_user_value = UserManage(name=str(sheet_value_list["credential"]),username=sheet_value_list["username"],password=aes_password,become_method=sheet_value_list["become_method"],
                                                become_user=sheet_value_list["become_user"],become_password=aes_become_password,public_key=sheet_value_list["public_key"])
            bulk_create_user_list.append(bulk_create_user_value)

            clint_host_list.append(copy.deepcopy(clint_host_dict))
        #批量插入主机管理
        logger.info('<---插入主机管理表 笔数:{0}--->'.format(len(bulk_create_host_list)))
        HostInfoManage.objects.bulk_create(bulk_create_host_list)

        logger.info('<---插入主机用户管理表 笔数:{0}--->'.format(len(bulk_create_user_list)))
        UserManage.objects.bulk_create(bulk_create_user_list)
        #后续进行连通性测试(异步)
        logger.info('<---接入主机连通性测试接口 (异步)--->')
        Clint_Host_Test(clint_host_list)

        data = {'code':1,'msg':'批量数据插入完成','data':'success'}
        return HttpResponse(json.dumps(data),content_type='application/json')
    except Exception as e:
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return HttpResponse(json.dumps(data),content_type='application/json')



#连通性检测
@Tr_async
def Clint_Host_Test(sheet_host_list):
    try:
        #sheet_host_list [{'hostip':'','port':22},{hostip':'','port':21}...]
        logger.info('<---目标主机连通性测试 参数:{0}--->'.format(sheet_host_list))
        #连通性测试
        clint_host_fai = []
        sheet_host_list_deep = copy.deepcopy(sheet_host_list)
        for host_info in sheet_host_list:
            sk = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
            sk.settimeout(1.5)
            try:
                sk.connect((host_info['hostip'],int(host_info['port'])))
                logger.info('<---目标主机连接成功 ip:{0} 端口:{1}--->'.format(host_info['hostip'],host_info['port']))
            except Exception as e:
                sheet_host_list_deep.remove(host_info)
                clint_host_fai.append(host_info)
    except Exception as e:
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
    else:
        logger.info('<---连接失败的数据:{0}--->'.format(clint_host_fai))
        if clint_host_fai:
            for clint_host_fai_upd in clint_host_fai:
                HostInfoManage.objects.filter(credential = clint_host_fai_upd['credential']).update(hoststatus=0)
        logger.info('<---需要ssh连通性测试的数据:{0}--->'.format(sheet_host_list_deep))
        if sheet_host_list_deep:
            Clint_SSH_Test(sheet_host_list_deep)
    finally:
        if sk:
            sk.close()


#ssh连接测试
def Clint_SSH_Test(sheet_host_list_deep):
    logger.info("<---启动校验SSH服务的接口--->")
    try:
        #获取凭证列表
        pz_list = [a['credential'] for a in sheet_host_list_deep]
        #获取凭证数据列表
        result = UserManage.objects.filter(name__in=pz_list)
        splic = []
        if result:
            #合并信息(将主机ip和该主机下的用户信息做对应)
            logger.info("<---合并主机信息和用户信息--->")
            for res in result:
                for shl in sheet_host_list_deep:
                    if str(res.name) == str(shl['credential']):
                        shl['result_pj']=res
                        splic.append(shl)

            #获取连接信息
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            #用于存储ssh连接失败的凭证id
            fai_ssh_host = []
            logger.info("<---准备ssh连接测试--->")
            for splic_data in splic:
                try:
                    if splic_data['result_pj'].become_method in ['su','enable']:
                        #密码解码
                        logger.info("<---su enable  用户密码解码--->")
                        pwd = AESEBC().Decrypt(splic_data['result_pj'].become_password)
                        ssh.connect(hostname=splic_data['hostip'], port=22, username=splic_data['result_pj'].become_user, password=pwd)
                        logger.info("<---主机:{0} ssh连接成功--->".format(splic_data['hostip']))
                    else:
                        #密码解码
                        logger.info("<---用户密码解码--->")
                        pwd = AESEBC().Decrypt(splic_data['result_pj'].password)
                        ssh.connect(hostname=splic_data['hostip'], port=22, username=splic_data['result_pj'].username, password=pwd)
                        logger.info("<---主机:{0} ssh连接成功--->".format(splic_data['hostip']))
                except Exception as ep:
                    fai_ssh_host.append(splic_data['credential'])
                    logger.info("<---ssh连接异常 异常信息:{0}  发生异常的行数:{1}--->".format(ep,ep.__traceback__.tb_lineno))

            #获取成功的凭证id
            suc_ssh_host = [suc for suc in pz_list if suc not in fai_ssh_host]

            logger.info("<---更新主机表状态 ssh连接失败的凭证id集合:{0}--->".format(fai_ssh_host))
            if fai_ssh_host:
                HostInfoManage.objects.filter(credential__in=fai_ssh_host).update(hoststatus=1)

            logger.info("<---更新主机表状态 ssh连接成功的凭证id集合:{0}--->".format(suc_ssh_host))
            if suc_ssh_host:
                HostInfoManage.objects.filter(credential__in=suc_ssh_host).update(hoststatus=2)

            logger.info("<---主机ssh连通性测试结束--->")
        else:
            logger.info("<---没有获取到凭证的数据--->")

    except Exception as e:
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
    finally:
        if ssh:
            ssh.close()




def Generate_DB_Excel(request):
    try:
        logger.info('<---调用文件生成接口--->')
        body = json.loads(request.body.decode())
        logger.info('<---接收的参数:{0}--->'.format(body))
        filepath = body.get('file_path','')

        if filepath:
            short_name, extension = os.path.splitext(filepath)
            if extension == '.xlsx':
                result_data = XLSX_Write_Excel(filepath)
            else:
                result_data = XLS_Write_Excel(filepath)

            logger.info("<---文件写入结束--->")
            return HttpResponse(json.dumps(result_data),content_type='application/json')
        else:
            logger.info("<---文件路径不能为空--->")
            data = {'code':2,'msg':"<---文件路径不能为空--->",'data':'Failure'}
            return HttpResponse(json.dumps(data),content_type='application/json')

    except Exception as e:
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return HttpResponse(json.dumps(data),content_type='application/json')


#读取xls后缀的excel文件
def XLS_Read_Excel(filepath):
    try:
        logger.info('<---调用读取excel.xls的文件接口>')
        #判断文件是否存在
        if os.path.isfile(filepath):
            #打开文件
            logger.info('<---打开文件--->')
            book = xlrd.open_workbook(filepath)
            sheet_host_list = []
            #获取页列表
            page_sheet = book.sheet_names()
            logger.info('<---获取所有页列表:{0}--->'.format(page_sheet))
            #所有页
            for sheet in page_sheet:
                #读取某一页
                sheet_dt = book.sheet_by_name(sheet)
                #行数
                rows = sheet_dt.nrows-1
                # cols = sheet_dt.ncols
                # im_data = np.zeros((rows,cols))
                logger.info('<---页名:{0} 数据行数:{1}--->'.format(sheet,rows))
                if rows > 0:
                    #获取第一行的key
                    sheet_host_key = sheet_dt.row_values(0)
                    #从第二行开始读 跳过说明行
                    for i in range(rows):
                        sheet_host_data = sheet_dt.row_values(i+1)
                        #形成字典对应关系  将float类型转换成int(xsls会把int 默认为float)
                        sheet_host_list.append(dict(zip(sheet_host_key,[a if isinstance(a,float)==False else int(a) for a in sheet_host_data ])))
                    return sheet_host_list
                else:
                    logger.info('路径文件:{} 无数据'.format(filepath))
                    data = {'code':2,'msg':'路径文件:{0} 表:{1} 无数据'.format(filepath,sheet),'data':'Failure'}
                    return HttpResponse(json.dumps(data),content_type='application/json')
        else:
            logger.info('路径文件:{} 没有找到文件'.format(filepath))
            data = {'code':2,'msg':'路径文件:{} 没有找到文件'.format(filepath),'data':'Failure'}
            return HttpResponse(json.dumps(data),content_type='application/json')
    except Exception as e:
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return HttpResponse(json.dumps(data),content_type='application/json')


#写入xls后缀的excel文件
def XLS_Write_Excel(filepath):
    try:
        logger.info('<---调用写入excel.xls的文件接口>')
        result = HostInfoManage.objects.all()
        if result:
            #创建一个excel对象
            book=xlwt.Workbook(encoding="utf-8",style_compression=0)
            # 创建一个sheet对象
            sheet = book.add_sheet('主机信息表', cell_overwrite_ok=True)
            #第一行的数据(列名)
            listing = ['hostname','hostgroup','hostip','port','credential','description','hoststatus']

            for i in range(0,len(listing)):
                sheet.write(0,i,listing[i])
            logger.info("<---列名写入完成: 列名 {0}-->".format(listing))
            logger.info("<---准备写入数据-->")
            for j in range(0,len(result)):
                result_data_list = [result[j].hostname,result[j].hostgroup,result[j].hostip,result[j].port,result[j].credential,result[j].description,result[j].hoststatus]
                for jr in range(0,len(result_data_list)):
                    sheet.write(j+1,jr,result_data_list[jr])
            logger.info("<---数据写入完成-->")
            book.save(filepath)
            logger.info("<---文件写入完成--->")
            data = {'code':1,'msg':"<---主机信息的excel文件生成完成--->",'data':'success'}
            return data
        else:
            logger.info("<---数据库表中无数据--->")
            data = {'code':2,'msg':"<---数据库表中无数据--->",'data':'Failure'}
            return data

    except Exception as e:
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return data


#读xlsx的excel文件
def XLSX_Read_Excel(filepath):
    try:
        logger.info('<---调用读取excel.xlsx的文件接口>')
        #判断文件是否存在
        if os.path.isfile(filepath):
            #打开文件
            logger.info('<---打开文件--->')
            workbook = openpyxl.load_workbook(filepath)
            sheet_host_key = []
            sheet_host_list = []
            #获取页列表
            page_sheet = workbook.sheetnames
            logger.info('<---获取所有页列表:{0}--->'.format(page_sheet))
            #所有页
            for sheet in page_sheet:
                #读取某一页
                sheet_dt = workbook[sheet]
                #行数
                rows = sheet_dt.max_row-1
                logger.info('<---页名:{0} 数据行数:{1}--->'.format(sheet,rows))
                if rows > 0:
                    #读文件
                    j = 0
                    for row_data in sheet_dt.rows:
                        if j == 0:
                            #第一行的key
                            sheet_host_key = [vs.value for vs in row_data]
                            j=1
                        else:
                            #形成字典对应关系  将float类型转换成int(xsls会把int 默认为float)
                            sheet_host_list.append(dict(zip(sheet_host_key,[a if isinstance(a,float)==False else int(a) for a in [vs.value for vs in row_data]])))
                    return sheet_host_list
                else:
                    logger.info('路径文件:{0} 表:{1} 无数据'.format(filepath,sheet))
                    data = {'code':2,'msg':'路径文件:{} 无数据'.format(filepath),'data':'Failure'}
                    return HttpResponse(json.dumps(data),content_type='application/json')
        else:
            logger.info('路径文件:{} 没有找到文件'.format(filepath))
            data = {'code':2,'msg':'路径文件:{} 没有找到文件'.format(filepath),'data':'Failure'}
            return HttpResponse(json.dumps(data),content_type='application/json')
    except Exception as e:
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return HttpResponse(json.dumps(data),content_type='application/json')



#生成xlsx的excel文件
def XLSX_Write_Excel(filepath):
    logger.info('<---生成主机表 excel.xlsx文件接口调用---->')
    try:
        result = HostInfoManage.objects.all()
        if result:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = '主机信息表'

            listing = ['hostname','hostgroup','hostip','port','credential','description','hoststatus']
            for i in range(0,len(listing)):
                worksheet.cell(1,i+1,listing[i])
            logger.info("<---列名写入完成: 列名 {0}-->".format(listing))
            logger.info("<---准备写入数据-->")
            for j in range(0,len(result)):
                result_data_list = [result[j].hostname,result[j].hostgroup,result[j].hostip,result[j].port,result[j].credential,result[j].description,result[j].hoststatus]
                for jr in range(0,len(result_data_list)):
                    worksheet.cell(j+2,jr+1,result_data_list[jr])
            logger.info("<---数据写入完成-->")
            workbook.save(filename=filepath)

            logger.info("<---文件写入完成--->")
            data = {'code':1,'msg':"<---主机信息的excel文件生成完成--->",'data':'success'}
            return data
        else:
            logger.info("<---数据库表中无数据--->")
            data = {'code':2,'msg':"<---数据库表中无数据--->",'data':'Failure'}
            return data
    except Exception as e:
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return data


#查询主机信息
def Sel_HostInfo(request):
    try:
        logger.info('<---调用查询主机信息接口  获取参数:{0}>'.format(request.GET))
        #获取页码
        page = request.GET.get('page',1)
        pagesize = int(request.GET.get('pagesize',15))
        hostname = request.GET.get('hostname','')
        hostip = request.GET.get('hostip','')
        credential = request.GET.get('credential','')
        result_list = []

        if hostname:
            sel_influence_result = HostInfoManage.objects.filter(hostname=hostname)
            if sel_influence_result.exists():
                result = Data_Pagination(page,pagesize,sel_influence_result)
            else:
                logger.info('<---没有找到主机名hostname:{0} 对应的数据-->'.format(hostname))
                data = {'code':2,'msg':'没有找到主机名hostname:{0} 对应的数据'.format(hostname),'data':'Failure'}
                return HttpResponse(json.dumps(data),content_type='application/json')
        elif credential:
            sel_influence_result = HostInfoManage.objects.filter(credential=credential)
            if sel_influence_result.exists():
                result = Data_Pagination(page,pagesize,sel_influence_result)
            else:
                logger.info('<---没有找到访问凭证credential:{0} 对应的数据--->'.format(credential))
                data = {'code':2,'msg':'没有找到访问凭证credential:{0} 对应的数据'.format(credential),'data':'Failure'}
                return HttpResponse(json.dumps(data),content_type='application/json')
        elif hostip:
            sel_influence_result = HostInfoManage.objects.filter(hostip=hostip)
            if sel_influence_result.exists():
                result = Data_Pagination(page,pagesize,sel_influence_result)
            else:
                logger.info('<---没有找到主机地址hostaddress:{0} 对应的数据--->'.format(hostip))
                data = {'code':2,'msg':'没有找到主机地址hostaddress:{0} 对应的数据'.format(hostip),'data':'Failure'}
                return HttpResponse(json.dumps(data),content_type='application/json')
        else:
            sel_influence_result = HostInfoManage.objects.all()
            if sel_influence_result.exists():
                result = Data_Pagination(page,pagesize,sel_influence_result)
            else:
                logger.info('<---没有找到数据信息--->')
                data = {'code':2,'msg':'没有找到数据信息','data':'Failure'}
                return HttpResponse(json.dumps(data),content_type='application/json')

        # logger.info('<---返回的分页数据:{}--->'.format(result))
        if result[0]:
            for result_sing in result[0]:
                result_list.append({'id':result_sing.id,'hostname':result_sing.hostname,'hostgroup':result_sing.hostgroup,'hostip':result_sing.hostip,
                                    'port':result_sing.port,'credential':result_sing.credential,'description':result_sing.description,'hoststatus':result_sing.hoststatus})
            data = {'code':1,'msg':'数据查询完成','total':result[1],'data':result_list}
        else:
            data = {'code':2,'msg':'已达最大页数','data':result[0]}
        logger.info('<---数据查询完成 返回通知:{0}--->'.format(data))
        return HttpResponse(json.dumps(data),content_type='application/json')
    except Exception as e :
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return HttpResponse(json.dumps(data),content_type='application/json')


#数据分页
def Data_Pagination(page,pagesize,result):
    logger.info('<---调用数据分页---->')
    paginator = Paginator(result,pagesize)
    try:
        if int(page) == 0 : page =1
        result_pag = paginator.page(int(page))
    except PageNotAnInteger:
        result_pag = paginator.page(1)
    except EmptyPage:
        # result_pag = paginator.page(paginator.num_pages)
        result_pag = []
    except Exception as e :
        logger.info('<---数据分页接口异常 异常信息:{0} 异常发生代码行数:{1}--->'.format(e,e.__traceback__.tb_lineno))

    return result_pag,paginator.count


#删除主机信息
def Del_HostInfo(request):
    try:
        logger.info('<---调用删除主机信息接口---->')
        body = json.loads(request.body.decode())
        logger.info('<---接收的参数:{}---->'.format(body))
        id_list = body.get('id_list',[])
        if id_list:
            del_influence_num = HostInfoManage.objects.filter(id__in=id_list)
            if del_influence_num.exists():
                del_influence_num.delete()
                logger.info('<---主机数据删除完成--->')
                data = {'code':1,'msg':'主机数据删除完成','data':'success'}
                return HttpResponse(json.dumps(data),content_type='application/json')
            else:
                data = {'code':2,'msg':'没有找到列表中的数据','data':'Failure'}
                return HttpResponse(json.dumps(data),content_type='application/json')
        else:
            data = {'code':2,'msg':'主机id不能为空','data':'Failure'}
            return HttpResponse(json.dumps(data),content_type='application/json')

    except Exception as e :
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return HttpResponse(json.dumps(data),content_type='application/json')

#修改主机信息
def Upd_HostInfo(request):
    logger.info('<---修改主机信息接口调用---->')
    body = json.loads(request.body.decode())
    try:
        logger.info('<---接收的参数:{}---->'.format(body))
        id = body.get('id','')
        hostname = body.get('hostname','')
        hostgroup = body.get('hostgroup','')
        hostip = body.get('hostip','')
        port = body.get('port','')
        credential = body.get('credential','')
        description = body.get('description','')
        hoststatus = body.get('hoststatus','')
        if id != '':
            sel_influence_num = HostInfoManage.objects.filter(id=int(id))
            if sel_influence_num.exists():
                upd_influence_num = sel_influence_num.update(hostname=hostname,hostgroup=hostgroup,hostip=hostip,port=int(port),
                                                             credential=credential,description=description,hoststatus=int(hoststatus))
                if upd_influence_num == 1:
                    logger.info('<---id:{0}的数据更新成功--->'.format(id))
                    data = {'code':1,'msg':'id:{0}的数据更新成功'.format(id),'data':'success'}
                    return HttpResponse(json.dumps(data),content_type='application/json')
                else:
                    data = {'code':2,'msg':'id:{0}的数据库更新失败'.format(id),'data':'Failure'}
                    return HttpResponse(json.dumps(data),content_type='application/json')
            else:
                data = {'code':2,'msg':'没有找到id:{0}的数据'.format(id),'data':'Failure'}
                return HttpResponse(json.dumps(data),content_type='application/json')
        else:
            data = {'code':2,'msg':'对比主机id不能为空','data':'Failure'}
            return HttpResponse(json.dumps(data),content_type='application/json')
    except Exception as e :
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return HttpResponse(json.dumps(data),content_type='application/json')


#主机用户添加
def Ins_UserManage(request):
    try:
        logger.info('<---调用主机用户添加接口--->')
        body = json.loads(request.body.decode())
        logger.info('<---接收的参数:{0}--->'.format(body))

        name = str(body.get('name',''))
        username = body.get('username','')
        password = body.get('password','')
        if password:
            password = AESEBC().Encrypt(password)
        become_method = body.get('become_method','')
        become_user = body.get('become_user','')
        become_password = body.get('become_password','')
        if become_password:
            become_password = AESEBC().Encrypt(become_password)
        public_key = body.get('public_key','')



        usermanage = UserManage()
        usermanage.name = name
        usermanage.username = username
        usermanage.password = password
        usermanage.become_method = become_method
        usermanage.become_user = become_user
        usermanage.become_password = become_password
        usermanage.public_key = public_key
        usermanage.save()

        logger.info('<---主机用户信息添加成功--->'.format(body))
        data = {'code':1,'msg':'主机用户信息添加成功','data':'success'}
        return HttpResponse(json.dumps(data),content_type='application/json')
    except Exception as e:
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return HttpResponse(json.dumps(data),content_type='application/json')



#删除主机用户信息
def Del_UserManage(request):
    try:
        logger.info('<---调用删除主机用户信息接口---->')
        body = json.loads(request.body.decode())
        logger.info('<---接收的参数:{}---->'.format(body))
        id_list = body.get('id_list',[])
        if id_list:
            del_influence_num = UserManage.objects.filter(id__in=id_list)
            if del_influence_num.exists():
                del_influence_num.delete()
                logger.info('<---主机用户数据删除完成--->')
                data = {'code':1,'msg':'主机用户数据删除完成','data':'success'}
                return HttpResponse(json.dumps(data),content_type='application/json')
            else:
                data = {'code':2,'msg':'没有找到列表中的数据','data':'Failure'}
                return HttpResponse(json.dumps(data),content_type='application/json')
        else:
            data = {'code':2,'msg':'主机用户id列表不能为空','data':'Failure'}
            return HttpResponse(json.dumps(data),content_type='application/json')

    except Exception as e :
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return HttpResponse(json.dumps(data),content_type='application/json')



#修改主机信息
def Upd_UserManage(request):
    logger.info('<---修改主机用户信息接口调用---->')
    body = json.loads(request.body.decode())
    try:
        logger.info('<---接收的参数:{}---->'.format(body))
        id = body.get('id','')
        name = str(body.get('name',''))
        username = body.get('username','')
        password = body.get('password','')
        become_method = body.get('become_method','')
        become_user = body.get('become_user','')
        become_password = body.get('become_password','')
        public_key = body.get('public_key','')

        if id != '':
            sel_influence_num = UserManage.objects.filter(id=int(id))
            if sel_influence_num.exists():
                upd_influence_num = sel_influence_num.update(name=name,username=username,password=password,become_method=become_method,
                                                             become_user=become_user,become_password=become_password,public_key=public_key)
                if upd_influence_num == 1:
                    logger.info('<---id:{0}的数据更新成功--->'.format(id))
                    data = {'code':1,'msg':'id:{0}的数据更新成功'.format(id),'data':'success'}
                    return HttpResponse(json.dumps(data),content_type='application/json')
                else:
                    data = {'code':2,'msg':'id:{0}的数据库更新失败'.format(id),'data':'Failure'}
                    return HttpResponse(json.dumps(data),content_type='application/json')
            else:
                data = {'code':2,'msg':'没有找到id:{0}的数据'.format(id),'data':'Failure'}
                return HttpResponse(json.dumps(data),content_type='application/json')
        else:
            data = {'code':2,'msg':'主机用户id不能为空','data':'Failure'}
            return HttpResponse(json.dumps(data),content_type='application/json')
    except Exception as e :
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return HttpResponse(json.dumps(data),content_type='application/json')



#查询主机用户信息
def Sel_UserManage(request):
    try:
        logger.info('<---调用查询主机用户信息接口  获取参数:{0}>'.format(request.GET))
        #获取页码
        page = request.GET.get('page',1)
        pagesize = int(request.GET.get('pagesize',15))
        name = str(request.GET.get('name',''))
        result_list = []

        if name:
            sel_influence_result = UserManage.objects.filter(name=name)
            if sel_influence_result.exists():
                result = Data_Pagination(page,pagesize,sel_influence_result)
            else:
                logger.info('<---没有找到凭证名称name:{0} 对应的数据-->'.format(name))
                data = {'code':2,'msg':'没有找到凭证名称name:{0} 对应的数据'.format(name),'data':'Failure'}
                return HttpResponse(json.dumps(data),content_type='application/json')
        else:
            sel_influence_result = UserManage.objects.all()
            if sel_influence_result.exists():
                result = Data_Pagination(page,pagesize,sel_influence_result)
            else:
                logger.info('<---没有找到数据信息--->')
                data = {'code':2,'msg':'没有找到数据信息','data':'Failure'}
                return HttpResponse(json.dumps(data),content_type='application/json')

        # logger.info('<---返回的分页数据:{}--->'.format(result))
        if result[0]:
            for result_sing in result[0]:
                result_list.append({'id':result_sing.id,'name':result_sing.name,'username':result_sing.username,'password':result_sing.password,'become_method':result_sing.become_method,
                                    'become_user':result_sing.become_user,'become_password':result_sing.become_password,'public_key':result_sing.public_key})
            data = {'code':1,'msg':'数据查询完成','total':result[1],'data':result_list}
        else:
            data = {'code':2,'msg':'已达最大页数','data':result[0]}
        logger.info('<---数据查询完成 返回通知:{0}--->'.format(data))
        return HttpResponse(json.dumps(data),content_type='application/json')
    except Exception as e :
        logger.info("<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno))
        data = {'code':2,'msg':"<---异常信息:{0}  发生异常的行数:{1}--->".format(e,e.__traceback__.tb_lineno),'data':'Failure'}
        return HttpResponse(json.dumps(data),content_type='application/json')

